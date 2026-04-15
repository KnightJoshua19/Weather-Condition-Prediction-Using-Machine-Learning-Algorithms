"""
Interactive Waste Management Policy ML Model
Enhanced with policymaker terminology and interactive features
For use with React.js frontend via API
"""

import os
import numpy as np
import pandas as pd
import warnings
import json
import datetime
warnings.filterwarnings('ignore')

from scipy.signal import butter, filtfilt
from sklearn.preprocessing import MinMaxScaler, StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, mean_squared_error, mean_absolute_error, r2_score
import matplotlib.pyplot as plt

# Excel Export Library
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    print("[LOADED] OpenPyXL library loaded successfully")
except ImportError:
    print("WARNING: openpyxl not installed. Install with: pip install openpyxl")

# Natural Language Generation Imports
try:
    from transformers import pipeline
    print("[LOADED] Transformers library loaded successfully")
except ImportError:
    print("WARNING: Transformers not installed. Install with: pip install transformers torch")
    
try:
    import nltk
    nltk.download('punkt', quiet=True)
    nltk.download('averaged_perceptron_tagger', quiet=True)
    from nltk.tokenize import sent_tokenize
    print("[LOADED] NLTK loaded successfully")
except ImportError:
    print("WARNING: NLTK not installed")

try:
    import xgboost as xgb
    print(f"[LOADED] XGBoost version {xgb.__version__} loaded successfully")
except ImportError:
    print("ERROR: XGBoost not installed. Run: pip install xgboost")
    raise

# ============================================================================
# POLICYMAKER-FOCUSED CONFIGURATION
# ============================================================================
class WasteManagementPolicyConfig:
    """
    Configuration for waste management policy analysis pipeline
    All terminology aligned with environmental policy standards
    """
    
    def __init__(self):
        self.data_path = r'C:\Users\Joshua\Waste-Management-Policymaking-Assistant\public_data_waste_fee.csv'
        
        # Preprocessing parameters
        self.filter_cutoff = 0.1
        self.filter_order = 4
        self.rolling_window = 5
        self.variance_threshold = 0.1
        
        # Model parameters
        self.test_size = 0.2
        self.random_state = 42
        self.epochs = 50
        self.batch_size = 32
        
        # Policy analysis parameters
        self.policy_assessment_threshold = 0.7
        
        # Feature mapping - policymaker-friendly names
        self.feature_labels = {
            'wden': 'Waste Density (kg/km²)',
            'urb': 'Urbanization Rate (%)',
            'msw': 'Municipal Solid Waste (tons/year)',
            'sor': 'Source Segregation Rate (%)',
            'fee': 'Fee Collection Mechanism',
            'organic': 'Organic Waste %',
            'paper': 'Paper/Cardboard %',
            'glass': 'Glass %',
            'plastic': 'Plastic %'
        }
        
        # Policy instruments mapping
        self.policy_instruments = {
            'segregation': 'Extended Producer Responsibility (EPR)',
            'fee': 'Pay-As-You-Throw (PAYT)',
            'collection': 'Multi-Stream Collection System (MSTS)',
            'organic': 'Organic Waste Diversion Mandate (OWDM)',
            'awareness': 'Environmental Awareness Program (EAP)'
        }
    
    def display_config(self):
        """Display configuration"""
        print("\n" + "="*70)
        print("WASTE MANAGEMENT POLICY ANALYSIS CONFIGURATION")
        print("="*70)
        for key, value in self.__dict__.items():
            if key != 'policy_instruments' and key != 'feature_labels':
                print(f"  {key:<30}: {value}")
        print("="*70 + "\n")


# ============================================================================
# ALGORITHM METRICS RECORDER
# ============================================================================
class AlgorithmMetricsRecorder:
    """
    Records algorithm performance metrics and complexity analysis
    Stores: Algorithm name, MSE, MAE, R-squared, Time/Space Complexity
    """
    
    def __init__(self, metrics_file='algorithm_metrics.json'):
        self.metrics_file = metrics_file
        self.metrics_history = self._load_existing_metrics()
    
    def _load_existing_metrics(self):
        """Load existing metrics from file if it exists"""
        if os.path.exists(self.metrics_file):
            try:
                with open(self.metrics_file, 'r') as f:
                    return json.load(f)
            except:
                return {'algorithms': []}
        return {'algorithms': []}
    
    def calculate_performance_metrics(self, y_true, y_pred, algorithm_name):
        """
        Calculate performance metrics: MSE, MAE, R-squared
        
        MSE (Mean Squared Error): Average of squared differences
            Formula: MSE = (1/n) * Σ(y_i - ŷ_i)²
        
        MAE (Mean Absolute Error): Average of absolute differences
            Formula: MAE = (1/n) * Σ|y_i - ŷ_i|
        
        R-squared (Coefficient of Determination): Proportion of variance explained
            Formula: R² = 1 - (SS_res / SS_tot)
            Range: [0, 1] where 1 is perfect prediction, 0 is no improvement
        """
        y_true = np.array(y_true).flatten()
        y_pred = np.array(y_pred).flatten()
        
        mse = mean_squared_error(y_true, y_pred)
        mae = mean_absolute_error(y_true, y_pred)
        r2 = r2_score(y_true, y_pred)
        
        metrics = {
            'algorithm': algorithm_name,
            'timestamp': datetime.datetime.now().isoformat(),
            'performance_metrics': {
                'MSE': {
                    'value': float(mse),
                    'formula': '(1/n) * Σ(y_i - ŷ_i)²',
                    'description': 'Mean Squared Error - penalizes larger errors'
                },
                'MAE': {
                    'value': float(mae),
                    'formula': '(1/n) * Σ|y_i - ŷ_i|',
                    'description': 'Mean Absolute Error - average prediction error'
                },
                'R_squared': {
                    'value': float(r2),
                    'value_range': '[0, 1]',
                    'formula': '1 - (SS_res / SS_tot)',
                    'description': 'Coefficient of Determination - proportion of variance explained',
                    'interpretation': self._interpret_r2(r2)
                }
            }
        }
        
        return metrics
    
    def _interpret_r2(self, r2_value):
        """Interpret R-squared value"""
        if r2_value >= 0.9:
            return "Excellent - Model explains >90% of variance"
        elif r2_value >= 0.7:
            return "Good - Model explains 70-90% of variance"
        elif r2_value >= 0.5:
            return "Moderate - Model explains 50-70% of variance"
        elif r2_value >= 0.3:
            return "Fair - Model explains 30-50% of variance"
        else:
            return "Poor - Model explains <30% of variance"
    
    def add_complexity_analysis(self, metrics, input_dim, hidden_layers, epochs, batch_size):
        """
        Add time and space complexity analysis to metrics
        
        Space Complexity: O(n * m) where n=input_dim, m=hidden_units
        Time Complexity: O(epochs * batch_size * parameters)
        """
        # Calculate total parameters
        total_params = 0
        layer_details = []
        
        for i in range(len(hidden_layers) - 1):
            if i == 0:
                params = input_dim * hidden_layers[i]
            else:
                params = hidden_layers[i-1] * hidden_layers[i]
            
            total_params += params
            layer_details.append({
                'layer': f'Layer_{i}',
                'input_units': input_dim if i == 0 else hidden_layers[i-1],
                'output_units': hidden_layers[i],
                'parameters': params
            })
        
        # Time complexity calculation
        ops_per_batch = sum([input_dim * hidden_layers[0]] + 
                           [hidden_layers[i-1] * hidden_layers[i] 
                            for i in range(1, len(hidden_layers))])
        total_ops = epochs * batch_size * ops_per_batch
        
        complexity_analysis = {
            'space_complexity': {
                'big_o': f'O(n*m)',
                'description': f'Linear with input dimensions and hidden units',
                'total_parameters': total_params,
                'layer_breakdown': layer_details
            },
            'time_complexity': {
                'big_o': f'O(e*b*ops)',
                'formula': f'epochs * batch_size * operations_per_batch',
                'values': {
                    'epochs': epochs,
                    'batch_size': batch_size,
                    'operations_per_forward_pass': ops_per_batch,
                    'total_estimated_operations': total_ops
                },
                'description': 'Proportional to training iterations and model size'
            }
        }
        
        metrics['complexity_analysis'] = complexity_analysis
        return metrics
    
    def save_metrics(self, metrics):
        """Save metrics to JSON file"""
        self.metrics_history['algorithms'].append(metrics)
        
        try:
            with open(self.metrics_file, 'w') as f:
                json.dump(self.metrics_history, f, indent=4)
            print(f"Metrics saved to {self.metrics_file}")
            return True
        except Exception as e:
            print(f"Error saving metrics: {e}")
            return False
    
    def get_latest_metrics(self):
        """Retrieve the latest recorded metrics"""
        if self.metrics_history['algorithms']:
            return self.metrics_history['algorithms'][-1]
        return None
    
    def get_metrics_summary(self):
        """Get summary of all recorded metrics"""
        summary = {
            'total_algorithms_recorded': len(self.metrics_history['algorithms']),
            'algorithms': []
        }
        
        for algo in self.metrics_history['algorithms']:
            algo_summary = {
                'algorithm': algo.get('algorithm'),
                'timestamp': algo.get('timestamp'),
                'mse': algo.get('performance_metrics', {}).get('MSE', {}).get('value'),
                'mae': algo.get('performance_metrics', {}).get('MAE', {}).get('value'),
                'r2': algo.get('performance_metrics', {}).get('R_squared', {}).get('value')
            }
            summary['algorithms'].append(algo_summary)
        
        return summary
    
    def print_metrics_report(self, metrics):
        """Print formatted metrics report"""
        print("\n" + "="*70)
        print("ALGORITHM PERFORMANCE METRICS REPORT")
        print("="*70)
        
        print(f"\nAlgorithm: {metrics.get('algorithm')}")
        print(f"Timestamp: {metrics.get('timestamp')}")
        
        perf = metrics.get('performance_metrics', {})
        print("\n--- Performance Metrics ---")
        print(f"MSE (Mean Squared Error): {perf.get('MSE', {}).get('value'):.6f}")
        print(f"  Formula: {perf.get('MSE', {}).get('formula')}")
        
        print(f"\nMAE (Mean Absolute Error): {perf.get('MAE', {}).get('value'):.6f}")
        print(f"  Formula: {perf.get('MAE', {}).get('formula')}")
        
        r2_info = perf.get('R_squared', {})
        print(f"\nR-squared: {r2_info.get('value'):.6f}")
        print(f"  Range: {r2_info.get('value_range')}")
        print(f"  Formula: {r2_info.get('formula')}")
        print(f"  Interpretation: {r2_info.get('interpretation')}")
        
        complexity = metrics.get('complexity_analysis', {})
        if complexity:
            print("\n--- Time & Space Complexity ---")
            space = complexity.get('space_complexity', {})
            print(f"Space Complexity: {space.get('big_o')}")
            print(f"  Total Parameters: {space.get('total_parameters'):,}")
            
            time = complexity.get('time_complexity', {})
            print(f"\nTime Complexity: {time.get('big_o')}")
            values = time.get('values', {})
            print(f"  Epochs: {values.get('epochs')}")
            print(f"  Batch Size: {values.get('batch_size')}")
            print(f"  Operations/Forward Pass: {values.get('operations_per_forward_pass'):,}")
            print(f"  Total Operations: {values.get('total_estimated_operations'):,}")
        
        print("\n" + "="*70 + "\n")
    
    def save_to_excel(self, excel_file='algorithm_metrics.xlsx'):
        """
        Save metrics to Excel file with minimalistic styling
        Includes performance metrics, time and space complexity
        """
        try:
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Algorithm Metrics"
            
            # Define minimalistic styles
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            header_font = Font(bold=True, size=11, color="000000")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Set column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 20
            
            row = 1
            
            # Add title
            ws[f'A{row}'] = "Algorithm Performance Metrics"
            ws[f'A{row}'].font = Font(bold=True, size=13)
            ws.merge_cells(f'A{row}:E{row}')
            row += 2
            
            # Add header row for performance metrics
            headers = ['Algorithm', 'MSE', 'MAE', 'R-squared', 'Timestamp']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = center_align
            
            row += 1
            
            # Add metrics data
            for algo in self.metrics_history['algorithms']:
                mse = algo.get('performance_metrics', {}).get('MSE', {}).get('value', 'N/A')
                mae = algo.get('performance_metrics', {}).get('MAE', {}).get('value', 'N/A')
                r2 = algo.get('performance_metrics', {}).get('R_squared', {}).get('value', 'N/A')
                
                ws[f'A{row}'] = algo.get('algorithm', '')
                ws[f'B{row}'] = mse if isinstance(mse, str) else f'{mse:.6f}'
                ws[f'C{row}'] = mae if isinstance(mae, str) else f'{mae:.6f}'
                ws[f'D{row}'] = r2 if isinstance(r2, str) else f'{r2:.6f}'
                ws[f'E{row}'] = algo.get('timestamp', '')
                
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    if col == 1 or col == 5:
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align
                
                row += 1
            
            # Add complexity analysis section
            row += 2
            ws[f'A{row}'] = "Time & Space Complexity Analysis"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
            
            # Complexity headers
            complexity_headers = ['Algorithm', 'Space Complexity', 'Time Complexity']
            for col_idx, header in enumerate(complexity_headers, 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = center_align
            
            row += 1
            
            # Add complexity data
            for algo in self.metrics_history['algorithms']:
                complexity = algo.get('complexity_analysis', {})
                space_big_o = complexity.get('space_complexity', {}).get('big_o', 'N/A')
                time_big_o = complexity.get('time_complexity', {}).get('big_o', 'N/A')
                
                ws[f'A{row}'] = algo.get('algorithm', '')
                ws[f'B{row}'] = space_big_o
                ws[f'C{row}'] = time_big_o
                
                for col in range(1, 4):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = center_align
                
                row += 1
            
            # Add detailed complexity metrics
            row += 2
            ws[f'A{row}'] = "Detailed Complexity Metrics"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            # Detailed headers
            detail_headers = ['Algorithm', 'Total Parameters', 'Epochs', 'Batch Size', 'Total Operations']
            for col_idx, header in enumerate(detail_headers, 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = center_align
            
            row += 1
            
            # Add detailed complexity data
            for algo in self.metrics_history['algorithms']:
                complexity = algo.get('complexity_analysis', {})
                total_params = complexity.get('space_complexity', {}).get('total_parameters', 'N/A')
                time_values = complexity.get('time_complexity', {}).get('values', {})
                epochs = time_values.get('epochs', 'N/A')
                batch_size = time_values.get('batch_size', 'N/A')
                total_ops = time_values.get('total_estimated_operations', 'N/A')
                
                ws[f'A{row}'] = algo.get('algorithm', '')
                ws[f'B{row}'] = total_params if isinstance(total_params, str) else f'{total_params:,}'
                ws[f'C{row}'] = epochs
                ws[f'D{row}'] = batch_size
                ws[f'E{row}'] = total_ops if isinstance(total_ops, str) else f'{total_ops:,}'
                
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = center_align
                
                row += 1
            
            # Save workbook
            wb.save(excel_file)
            print(f"Metrics exported to {excel_file}")
            return True
            
        except Exception as e:
            print(f"Error saving to Excel: {e}")
            return False
    
    def export_metrics_to_excel(self, excel_file='algorithm_metrics.xlsx'):
        """
        Convenience method to export metrics to Excel file
        Wraps save_to_excel method
        """
        return self.save_to_excel(excel_file)


# ============================================================================
# ENHANCED DATA PREPROCESSING WITH POLICY CONTEXT
# ============================================================================
class PolicyDataPreprocessor:
    """
    Data preprocessing with policy-specific handling
    Optimized for waste management decision support
    """
    
    def __init__(self, config):
        self.config = config
        self.scaler = MinMaxScaler()
        self.original_df = None
        self.processed_df = None
    
    def load_data(self):
        """Load waste management data"""
        print("[POLICY-DATA] Loading regional waste data...")
        self.original_df = pd.read_csv(self.config.data_path)
        print(f"  [LOADED] {len(self.original_df)} regional records loaded")
        return self.original_df
    
    def apply_filter(self, series, cutoff=0.1, fs=1.0, order=4):
        """Apply Butterworth filter for smooth data"""
        if len(series) > order:
            b, a = butter(order, cutoff / (0.5 * fs), btype='low')
            return pd.Series(filtfilt(b, a, series), index=series.index)
        return series
    
    def clean_data(self):
        """Clean data for policy analysis"""
        print("[POLICY-DATA] Cleaning regional datasets...")
        df = self.original_df.copy()
        
        numeric_cols = df.select_dtypes(include='number').columns
        for col in numeric_cols:
            df[col] = df[col].astype(float)
            df[col] = self.apply_filter(
                df[col],
                cutoff=self.config.filter_cutoff,
                order=self.config.filter_order
            )
        
        column_mean = df.select_dtypes(include="number").rolling(
            window=self.config.rolling_window,
            center=True
        ).mean()
        df = df.fillna(column_mean)
        df.update(column_mean)
        
        initial_rows = len(df)
        df = df.drop_duplicates()
        removed_rows = initial_rows - len(df)
        print(f"  [CLEANED] Data cleaned: {removed_rows} duplicates removed")
        
        self.processed_df = df
        return df
    
    def engineer_features(self):
        """Engineer policy-relevant features"""
        print("[POLICY-DATA] Engineering policy-relevant features...")
        df = self.processed_df.copy()
        
        def mean_absolute_difference(series):
            return np.mean(np.abs(series - np.mean(series)))
        
        removed_features = []
        numeric_cols = df.select_dtypes(include='number').columns
        for col in numeric_cols:
            mad = mean_absolute_difference(df[col])
            if mad < self.config.variance_threshold:
                removed_features.append(col)
                df.drop(columns=[col], inplace=True)
        
        if removed_features:
            print(f"  [REMOVED] Low-variance features removed: {len(removed_features)}")
        
        self.processed_df = df
        return df
    
    def scale_features(self):
        """Normalize features"""
        print("[POLICY-DATA] Normalizing feature scales...")
        df = self.processed_df.copy()
        
        numeric_cols = df.select_dtypes(include='number').columns
        df[numeric_cols] = self.scaler.fit_transform(df[numeric_cols])
        
        self.processed_df = df
        return df
    
    def get_processed_data(self):
        """Execute full preprocessing"""
        self.load_data()
        self.clean_data()
        self.engineer_features()
        self.scale_features()
        return self.processed_df


# ============================================================================
# INTERACTIVE POLICY-FOCUSED TEXT GENERATION
# ============================================================================
class PolicyIntelligenceGenerator:
    """
    Generates policy recommendations and insights
    Uses trained ML model to identify patterns
    Outputs policymaker-friendly reports
    """
    
    def __init__(self, config):
        self.config = config
        self.model = None
    
    # ============================================================================
    # MAIN ALGORITHM: NEURAL NETWORK ARCHITECTURE FOR POLICY ANALYSIS
    # ============================================================================
    # This section implements the core machine learning algorithm for waste management
    # policy analysis. XGBoost is used to process regional waste data
    # and generate policy recommendations based on learned patterns.
    #
    # Algorithm Overview:
    # - Input: Preprocessed regional waste management features (e.g., waste density,
    #   urbanization rate, segregation rates)
    # - Architecture: Gradient boosting with 100 estimators, max_depth=6
    # - Output: Binary classification for policy readiness assessment
    #
    # Time Complexity: O(n_trees * tree_depth * n_features * log(n_samples))
    # Space Complexity: O(n_trees) for tree structures
    #
    # Key Parameters:
    # 1. n_estimators: 100 boosting rounds
    # 2. max_depth: 6 for balanced model complexity
    # 3. learning_rate: 0.1 for regularization
    # 4. subsample: 0.8 for row subsampling
    # 5. colsample_bytree: 0.8 for column subsampling
    #
    # Advantages: Better generalization on structured data, faster inference
    # ============================================================================
    def build_model(self, input_dim):
        """Build XGBoost policy analysis classifier"""
        print("[POLICY-MODEL] Building XGBoost policy analysis model...")
        
        self.model = xgb.XGBClassifier(
            n_estimators=100,
            max_depth=6,
            learning_rate=0.1,
            subsample=0.8,
            colsample_bytree=0.8,
            random_state=42,
            eval_metric='logloss',
            verbosity=0,
            objective='binary:logistic'
        )
        
        print(f"  [CREATED] XGBoost policy analysis model initialized")
        return self.model
    
    def train_model(self, X_train, y_train, X_val, y_val):
        """Train XGBoost model"""
        print("[POLICY-MODEL] Training XGBoost policy model...")
        
        # Train XGBoost with early stopping
        self.model.fit(
            X_train, y_train,
            eval_set=[(X_val, y_val)],
            early_stopping_rounds=10,
            verbose=False
        )
        
        # Calculate validation accuracy
        val_accuracy = self.model.score(X_val, y_val)
        print(f"  [COMPLETED] Model training complete: {val_accuracy:.2%} accuracy")
        
        # Return a simple history-like object for compatibility
        history = {
            'train_accuracy': [self.model.score(X_train, y_train)],
            'val_accuracy': [val_accuracy]
        }
        return history
    
    def visualize_complexity(self, input_sizes=None):
        """
        Visualize time and space complexity of the XGBoost algorithm
        
        Time Complexity: O(n_trees * tree_depth * n_features * log(n_samples))
        Space Complexity: O(n_trees) for tree structures
        
        This method plots complexity curves for different input dimensions
        """
        if input_sizes is None:
            input_sizes = [10, 20, 50, 100, 200, 500]
        
        # Calculate space complexity (number of trees)
        n_trees = 100  # XGBoost estimators
        space_complexity = [n_trees * dim for dim in input_sizes]  # Rough estimate
        
        # Calculate time complexity (rough estimate)
        tree_depth = 6
        sample_size = 1000  # Approximate training samples
        time_complexity = []
        for dim in input_sizes:
            # Time ≈ n_trees * tree_depth * dim * log(sample_size)
            operations = n_trees * tree_depth * dim * np.log(sample_size + 1)
            time_complexity.append(operations)
        
        # Create visualization
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
        
        # Space complexity plot
        ax1.plot(input_sizes, space_complexity, 'b-o', linewidth=2, markersize=6)
        ax1.set_title('Space Complexity: O(n_trees × features)')
        ax1.set_xlabel('Input Dimension')
        ax1.set_ylabel('Estimated Tree Size')
        ax1.grid(True, alpha=0.3)
        
        # Time complexity plot
        ax2.plot(input_sizes, time_complexity, 'r-s', linewidth=2, markersize=6)
        ax2.set_title('Time Complexity: O(n_trees × depth × features × log(n))')
        ax2.set_xlabel('Input Dimension')
        ax2.set_ylabel('Estimated Operations')
        ax2.grid(True, alpha=0.3)
        
        plt.tight_layout()
        plt.savefig('complexity_analysis.png', dpi=300, bbox_inches='tight')
        print("[SAVED] Complexity visualization saved as 'complexity_analysis.png'")
        plt.show()
    
    def generate_policy_brief(self, region_data):
        """
        Generate executive policy brief for a region
        Input: Regional waste management data
        Output: Policy recommendations
        """
        brief = {
            'region': region_data.get('region', 'Unknown'),
            'assessment_date': pd.Timestamp.now().isoformat(),
            'executive_summary': self._create_executive_summary(region_data),
            'key_findings': self._analyze_findings(region_data),
            'policy_recommendations': self._recommend_policies(region_data),
            'implementation_roadmap': self._create_roadmap(region_data)
        }
        return brief
    
    def _create_executive_summary(self, data):
        """Create executive summary"""
        return f"""
        Regional Waste Management Policy Assessment for {data.get('region')}
        
        This assessment evaluates current waste management practices against 
        international and Philippine environmental standards. Recommendations are 
        based on data-driven analysis of key performance indicators.
        """
    
    def _analyze_findings(self, data):
        """Analyze key findings"""
        findings = []
        
        waste_density = float(data.get('wden', 0))
        sorting_rate = float(data.get('sor', 0))
        urbanization = float(data.get('urb', 0))
        
        if waste_density > 300000:
            findings.append({
                'indicator': 'Waste Generation Pressure',
                'status': 'CRITICAL',
                'finding': f'Waste density of {waste_density:,.0f} kg/km² indicates severe pressure',
                'implication': 'Immediate infrastructure scaling required'
            })
        
        if sorting_rate < 30:
            findings.append({
                'indicator': 'Waste Segregation Effectiveness',
                'status': 'CRITICAL',
                'finding': f'Sorting rate of {sorting_rate:.1f}% is below 50% threshold',
                'implication': 'Enhanced segregation programs required'
            })
        
        if urbanization > 70:
            findings.append({
                'indicator': 'Urban Concentration',
                'status': 'HIGH',
                'finding': f'{urbanization:.1f}% urbanization rate indicates concentrated population',
                'implication': 'Urban-focused policy needed'
            })
        
        return findings
    
    def _recommend_policies(self, data):
        """Generate policy recommendations"""
        recommendations = []
        
        sorting_rate = float(data.get('sor', 0))
        urbanization = float(data.get('urb', 0))
        organic_pct = float(data.get('organic', 0))
        
        if sorting_rate < 50:
            recommendations.append({
                'policy': 'Waste Segregation Strengthening',
                'instrument': 'Waste Segregation Ordinance',
                'priority': 'CRITICAL',
                'actions': [
                    'Establish barangay Material Recovery Facilities (MRF)',
                    'Launch "Zero Waste Community" program',
                    'Implement color-coded bins citywide'
                ],
                'expected_impact': 'Increase sorting rate to 70%+ within 12 months'
            })
        
        if organic_pct > 50:
            recommendations.append({
                'policy': 'Organic Waste Management',
                'instrument': 'Organic Waste Diversion Mandate',
                'priority': 'HIGH',
                'actions': [
                    'Establish composting facilities',
                    'Promote home composting programs',
                    'Partner with agricultural institutions'
                ],
                'expected_impact': 'Divert 40%+ of MSW from landfills'
            })
        
        if urbanization > 60:
            recommendations.append({
                'policy': 'Urban Waste Management System',
                'instrument': 'Multi-Stream Collection System (MSTS)',
                'priority': 'HIGH',
                'actions': [
                    'Establish dedicated collection routes',
                    'Deploy collection vehicles for segregated streams',
                    'Implement real-time tracking system'
                ],
                'expected_impact': 'Improve collection efficiency to 90%+'
            })
        
        return recommendations
    
    def _create_roadmap(self, data):
        """Create implementation roadmap"""
        return {
            'phase_1': {
                'duration': '0-3 months',
                'focus': 'Policy Formulation & Stakeholder Engagement',
                'activities': [
                    'Conduct stakeholder consultations',
                    'Finalize policy framework',
                    'Secure budget allocation'
                ]
            },
            'phase_2': {
                'duration': '3-6 months',
                'focus': 'Infrastructure Development',
                'activities': [
                    'Construct Material Recovery Facilities',
                    'Procure collection equipment',
                    'Train waste management personnel'
                ]
            },
            'phase_3': {
                'duration': '6-12 months',
                'focus': 'Implementation & Monitoring',
                'activities': [
                    'Launch waste segregation program',
                    'Monitor policy compliance',
                    'Promote public awareness'
                ]
            }
        }
    
    def evaluate_region(self, region_data):
        """Evaluate a region's policy readiness"""
        evaluation = {
            'region': region_data.get('region'),
            'waste_management_score': self._calculate_wmm_score(region_data),
            'policy_readiness': self._assess_policy_readiness(region_data),
            'priority_interventions': self._identify_priority_interventions(region_data)
        }
        return evaluation
    
    def _calculate_wmm_score(self, data):
        """Calculate Waste Management Maturity score (0-100)"""
        weights = {
            'sor': 0.25,  # Sorting rate
            'urb': 0.15,  # Urbanization
            'msw': 0.15   # MSW handling
        }
        
        score = (
            float(data.get('sor', 0)) * weights['sor'] +
            float(data.get('urb', 0)) * weights['urb'] +
            min(float(data.get('msw', 0)) / 100, 100) * weights['msw']
        )
        return round(score, 2)
    
    def _assess_policy_readiness(self, data):
        """Assess policy readiness level"""
        wmm_score = self._calculate_wmm_score(data)
        
        if wmm_score < 30:
            return 'CRITICAL - Immediate intervention required'
        elif wmm_score < 50:
            return 'INADEQUATE - Significant improvements needed'
        elif wmm_score < 70:
            return 'DEVELOPING - On track for improvement'
        else:
            return 'ADVANCED - Strong policy implementation'
    
    def _identify_priority_interventions(self, data):
        """Identify priority interventions"""
        interventions = []
        
        if float(data.get('sor', 0)) < 40:
            interventions.append({
                'intervention': 'Waste Segregation Program',
                'urgency': 'CRITICAL'
            })
        
        if float(data.get('organic', 0)) > 50:
            interventions.append({
                'intervention': 'Organic Waste Treatment',
                'urgency': 'HIGH'
            })
        
        return interventions


# ============================================================================
# EXTRACTIVE NATURAL LANGUAGE GENERATION FOR DATA-TO-TEXT
# ============================================================================
class ExtractiveNLGEngine:
    """
    Extractive Natural Language Generation Engine
    Generates policy responses using pre-trained transformers
    Employs extractive summarization for data-to-text tasks
    """
    
    def __init__(self):
        """Initialize NLG models"""
        self.summarizer = None
        self.qa_pipeline = None
        self._initialize_models()
    
    def _initialize_models(self):
        """Initialize pre-trained transformer models"""
        try:
            # Load pre-trained BART summarizer for extractive summarization
            print("[LOADING] Initializing BART summarization model...")
            self.summarizer = pipeline(
                "summarization",
                model="facebook/bart-large-cnn",
                device=-1  # Use CPU, change to 0 for GPU
            )
            print("[SUCCESS] BART extractive summarization model loaded successfully")
        except Exception as e:
            print(f"[WARNING] BART model download/initialization issue ({e})")
            print("[FALLBACK] Using rule-based extractive NLG approach")
            self.summarizer = None
    
    def extract_key_sentences(self, text_block, num_sentences=3):
        """
        Extract key sentences from text using simple scoring
        Extractive approach: select most important existing sentences
        """
        try:
            sentences = sent_tokenize(text_block)
        except:
            sentences = text_block.split('. ')
        
        if len(sentences) <= num_sentences:
            return ' '.join(sentences)
        
        # Simple extractive scoring based on length and content
        scored_sentences = []
        important_keywords = ['critical', 'high', 'priority', 'improve', 'urgent', 'need', 'require']
        
        for sent in sentences:
            score = len(sent.split())  # Longer sentences score higher
            # Boost score for sentences with important keywords
            score += sum(10 for kw in important_keywords if kw in sent.lower())
            scored_sentences.append((sent, score))
        
        # Sort by score and extract top sentences
        top_sentences = sorted(scored_sentences, key=lambda x: x[1], reverse=True)[:num_sentences]
        # Return in original order
        top_sentences = sorted(top_sentences, key=lambda x: sentences.index(x[0]))
        
        return ' '.join([s[0] for s in top_sentences])
    
    def generate_data_to_text(self, data_dict, question_type):
        """
        Generate natural language text from structured data
        Uses extractive approach: combines key facts from data
        """
        if question_type == "improvements":
            return self._generate_improvements_text(data_dict)
        elif question_type == "recommendations":
            return self._generate_recommendations_text(data_dict)
        elif question_type == "analysis":
            return self._generate_analysis_text(data_dict)
        else:
            return self._generate_summary_text(data_dict)
    
    def _generate_improvements_text(self, data):
        """Generate extractive text for improvement needs"""
        facts = []
        
        if data.get('low_segregation'):
            facts.append("Low waste segregation rates require immediate attention through community education and infrastructure improvements.")
        
        if data.get('high_urbanization'):
            facts.append("High urbanization levels demand multi-stream collection systems and dedicated waste management infrastructure.")
        
        if data.get('high_organic'):
            facts.append("High organic waste composition presents significant opportunity for composting programs and diversion from landfills.")
        
        if not facts:
            facts.append("Current waste management practices require assessment for optimal policy implementation.")
        
        return self._extract_and_format(facts)
    
    def _generate_recommendations_text(self, data):
        """Generate extractive text for policy recommendations"""
        recommendations = []
        
        if data.get('segregation_priority'):
            recommendations.append("Establishing waste segregation ordinances with barangay Material Recovery Facilities will significantly improve source separation rates.")
        
        if data.get('organic_priority'):
            recommendations.append("Community composting programs and agricultural partnerships can effectively divert organic waste from landfill disposal.")
        
        if data.get('urban_priority'):
            recommendations.append("Implementing integrated urban waste management systems with real-time tracking ensures efficient collection and processing.")
        
        recommendations.append("Public awareness campaigns focusing on environmental stewardship will strengthen policy compliance and community participation.")
        
        return self._extract_and_format(recommendations[:3])
    
    def _generate_analysis_text(self, data):
        """Generate extractive text for data analysis"""
        insights = []
        
        if data.get('regions_critical'):
            insights.append(f"Analysis reveals {data.get('regions_critical', 0)} regions in critical status requiring immediate policy intervention and resource allocation.")
        
        if data.get('avg_metric'):
            insights.append(f"Average performance metrics show {data.get('avg_metric', 'areas')} for optimization and targeted improvement initiatives.")
        
        if data.get('trend'):
            insights.append(f"Current trends indicate {data.get('trend', 'opportunities')} for policy enhancement and infrastructure development.")
        
        return self._extract_and_format(insights)
    
    def _generate_summary_text(self, data):
        """Generate extractive summary text"""
        summary = []
        
        if data.get('total_regions'):
            summary.append(f"Regional assessment covers {data.get('total_regions')} areas across multiple waste management indicators.")
        
        if data.get('key_metrics'):
            summary.append(f"Key metrics include segregation rates, urbanization levels, and organic waste composition analysis.")
        
        summary.append("Data-driven insights support evidence-based policy formulation and implementation strategies.")
        
        return self._extract_and_format(summary)
    
    def _extract_and_format(self, statements):
        """Format extracted statements into coherent response"""
        result = []
        for stmt in statements:
            if stmt and len(stmt.strip()) > 0:
                result.append(stmt.strip())
        
        return " ".join(result)
    
    def calculate_complexity_metrics(self, input_dim, hidden_layers=[64, 32, 16, 1]):
        """
        Calculate space and time complexity of NLG + ML pipeline
        Returns detailed complexity analysis
        """
        # Space Complexity: O(input_dim * hidden_1 + hidden_1 * hidden_2 + ... + hidden_n * output)
        total_params = 0
        for i in range(len(hidden_layers) - 1):
            if i == 0:
                total_params += input_dim * hidden_layers[i]
            else:
                total_params += hidden_layers[i-1] * hidden_layers[i]
        
        total_params += hidden_layers[-2] * hidden_layers[-1]  # Final layer
        
        # Add NLG model parameters (BART-large-cnn: ~406M parameters)
        nlg_params = 406000000  # BART-large-cnn model size
        
        # Time Complexity calculation
        epochs = 50
        batch_size = 32
        time_ops = epochs * batch_size * (input_dim * 64 + 64*32 + 32*16 + 16*1)
        
        return {
            'space_complexity': f"O(n*m) where n={input_dim}, total parameters={total_params:,}",
            'total_parameters': total_params,
            'nlg_model_parameters': nlg_params,
            'time_complexity': f"O(e*b*ops) where e={epochs}, b={batch_size}, ops={input_dim*64 + 64*32 + 32*16 + 16*1:,}",
            'estimated_operations': time_ops
        }
    
    def demonstrate_bart_extraction(self, long_text, max_length=100, min_length=50):
        """
        Demonstrate BART-based extractive summarization
        BART (Bidirectional and Auto-Regressive Transformers) with facebook/bart-large-cnn
        performs extractive and abstractive summarization
        """
        if self.summarizer is None:
            return "[FALLBACK] BART model not available. Using rule-based extraction."
        
        try:
            print("[BART] Processing text with facebook/bart-large-cnn model...")
            
            # Summarize using BART
            summary = self.summarizer(
                long_text,
                max_length=max_length,
                min_length=min_length,
                do_sample=False
            )
            
            return summary[0]['summary_text']
        except Exception as e:
            print(f"[BART_ERROR] {e}")
            return self.extract_key_sentences(long_text)


# ============================================================================
# INTERACTIVE QUESTION-ANSWERING SYSTEM
# ============================================================================
class PolicyQuestionAnswering:
    """
    Interactive system for answering policy-related questions
    Based solely on the provided waste management data
    Enhanced with extractive NLG for improved responses
    """
    
    def __init__(self, data, generator):
        self.data = data
        self.generator = generator
        self.numeric_data = data.select_dtypes(include='number')
        self.nlg_engine = ExtractiveNLGEngine()  # Initialize NLG engine
    
    def answer_question(self, question):
        """Answer user questions based on data analysis with NLG enhancement"""
        question = question.lower().strip()
        
        # Route to appropriate analysis method
        if 'improve' in question or 'improvement' in question:
            return self._analyze_improvements_needed()
        elif 'region' in question and ('need' in question or 'worst' in question):
            return self._identify_regions_needing_attention()
        elif 'segregation' in question or 'sorting' in question:
            return self._analyze_segregation_issues()
        elif 'organic' in question or 'compost' in question:
            return self._analyze_organic_waste_issues()
        elif 'urban' in question or 'city' in question:
            return self._analyze_urban_challenges()
        elif 'score' in question or 'maturity' in question:
            return self._calculate_overall_maturity()
        elif 'recommend' in question or 'policy' in question:
            return self._provide_policy_recommendations()
        elif 'complex' in question or 'algorithm' in question:
            return self._analyze_algorithm_complexity()
        else:
            return self._provide_general_summary()
    
    def _analyze_improvements_needed(self):
        """Analyze what improvements are needed using NLG"""
        data = {
            'low_segregation': self.numeric_data['sor'].mean() < 50,
            'high_urbanization': len(self.numeric_data[self.numeric_data['urb'] > 70]) > 0,
            'high_organic': len(self.numeric_data[self.numeric_data['organic'] > 50]) > 0
        }
        
        # Get traditional analysis
        analysis = []
        avg_sor = self.numeric_data['sor'].mean()
        low_sor_regions = len(self.numeric_data[self.numeric_data['sor'] < 30])
        analysis.append(f"Segregation Rate: Average {avg_sor:.1f}%, {low_sor_regions} regions below 30% threshold")
        
        high_urb_regions = len(self.numeric_data[self.numeric_data['urb'] > 70])
        analysis.append(f"Urban Pressure: {high_urb_regions} regions with >70% urbanization rate")
        
        high_organic_regions = len(self.numeric_data[self.numeric_data['organic'] > 50])
        analysis.append(f"Organic Waste: {high_organic_regions} regions with >50% organic content")
        
        # Generate NLG response
        nlg_response = self.nlg_engine.generate_data_to_text(data, "improvements")
        
        return f"[ANALYSIS] Key Improvements Needed:\n\n{nlg_response}\n\n[METRICS] Detailed Statistics:\n" + "\n".join(f"• {item}" for item in analysis)
    
    def _identify_regions_needing_attention(self):
        """Identify regions with the most critical needs using NLG"""
        # Calculate maturity scores for all regions
        scores = []
        for idx, row in self.data.iterrows():
            region_data = row.to_dict()
            score = self.generator._calculate_wmm_score(region_data)
            scores.append((region_data.get('region', f'Region {idx}'), score))
        
        # Sort by score (lowest first)
        scores.sort(key=lambda x: x[1])
        critical_regions = [f"{region}: {score:.1f}/100" for region, score in scores[:5]]
        
        nlg_data = {
            'regions_critical': len([s for s in scores if s[1] < 30]),
            'avg_metric': f"maturity scores of {sum(s[1] for s in scores)/len(scores):.1f}/100"
        }
        
        nlg_response = self.nlg_engine.generate_data_to_text(nlg_data, "analysis")
        
        return f"[PRIORITY] Critical Attention Required:\n\n{nlg_response}\n\n[REGIONS] Top Priority Areas:\n" + "\n".join(f"• {region}" for region in critical_regions)
    
    def _analyze_segregation_issues(self):
        """Analyze waste segregation challenges using NLG"""
        sor_stats = self.numeric_data['sor'].describe()
        
        data = {
            'segregation_priority': sor_stats['mean'] < 50,
            'low_regions': len(self.numeric_data[self.numeric_data['sor'] < 30])
        }
        
        issues = [
            f"Average segregation rate: {sor_stats['mean']:.1f}%",
            f"Lowest rate: {sor_stats['min']:.1f}%",
            f"Highest rate: {sor_stats['max']:.1f}%",
            f"Regions below 30%: {len(self.numeric_data[self.numeric_data['sor'] < 30])}"
        ]
        
        nlg_response = self.nlg_engine.generate_data_to_text(data, "recommendations")
        
        return f"[SEGREGATION] Waste Segregation Analysis:\n\n{nlg_response}\n\n[STATUS] Current Metrics:\n" + "\n".join(f"• {item}" for item in issues)
    
    def _analyze_organic_waste_issues(self):
        """Analyze organic waste management issues using NLG"""
        organic_stats = self.numeric_data['organic'].describe()
        
        data = {
            'organic_priority': organic_stats['mean'] > 40,
            'high_regions': len(self.numeric_data[self.numeric_data['organic'] > 50])
        }
        
        issues = [
            f"Average organic waste: {organic_stats['mean']:.1f}% of total waste",
            f"Regions with >50% organic: {len(self.numeric_data[self.numeric_data['organic'] > 50])}",
            f"Potential for composting: High in {len(self.numeric_data[self.numeric_data['organic'] > 40])} regions"
        ]
        
        nlg_response = self.nlg_engine.generate_data_to_text(data, "recommendations")
        
        return f"[ORGANIC] Organic Waste Management:\n\n{nlg_response}\n\n[ANALYSIS] Key Insights:\n" + "\n".join(f"• {item}" for item in issues)
    
    def _analyze_urban_challenges(self):
        """Analyze urban waste management challenges using NLG"""
        urb_stats = self.numeric_data['urb'].describe()
        
        data = {
            'urban_priority': urb_stats['mean'] > 60,
            'high_density_regions': len(self.numeric_data[self.numeric_data['urb'] > 70])
        }
        
        issues = [
            f"Average urbanization: {urb_stats['mean']:.1f}%",
            f"Highly urbanized regions (>70%): {len(self.numeric_data[self.numeric_data['urb'] > 70])}",
            f"Urban waste density correlation: Higher collection challenges in dense areas"
        ]
        
        nlg_response = self.nlg_engine.generate_data_to_text(data, "recommendations")
        
        return f"[URBAN] Urban Waste Management Challenges:\n\n{nlg_response}\n\n[DATA] Statistical Summary:\n" + "\n".join(f"• {item}" for item in issues)
    
    def _calculate_overall_maturity(self):
        """Calculate overall waste management maturity using NLG"""
        scores = []
        for idx, row in self.data.iterrows():
            region_data = row.to_dict()
            score = self.generator._calculate_wmm_score(region_data)
            scores.append(score)
        
        avg_score = sum(scores) / len(scores) if scores else 0
        distribution = {
            'Critical (<30)': len([s for s in scores if s < 30]),
            'Inadequate (30-50)': len([s for s in scores if 30 <= s < 50]),
            'Developing (50-70)': len([s for s in scores if 50 <= s < 70]),
            'Advanced (>70)': len([s for s in scores if s >= 70])
        }
        
        nlg_data = {
            'total_regions': len(scores),
            'avg_metric': f"maturity score of {avg_score:.1f}/100"
        }
        
        nlg_response = self.nlg_engine.generate_data_to_text(nlg_data, "summary")
        
        result = f"[ASSESSMENT] Waste Management Maturity Review:\n\n{nlg_response}\n\n[DISTRIBUTION] Score Breakdown:\n"
        for level, count in distribution.items():
            result += f"• {level}: {count} regions\n"
        
        return result
    
    def _provide_policy_recommendations(self):
        """Provide general policy recommendations using NLG"""
        data = {
            'segregation_priority': self.numeric_data['sor'].mean() < 50,
            'organic_priority': self.numeric_data['organic'].mean() > 40,
            'urban_priority': self.numeric_data['urb'].mean() > 60
        }
        
        nlg_response = self.nlg_engine.generate_data_to_text(data, "recommendations")
        
        return f"[RECOMMENDATIONS] Evidence-Based Policy Suggestions:\n\n{nlg_response}"
    
    def _analyze_algorithm_complexity(self):
        """Analyze and visualize algorithm complexity"""
        input_dims = [10, 20, 50, 100, 200, 500]
        
        # Calculate complexity metrics
        space_complexity = []
        time_complexity = []
        
        for dim in input_dims:
            params = dim * 64 + 64 * 32 + 32 * 16 + 16 * 1
            space_complexity.append(params)
            
            ops = 50 * 32 * (dim * 64 + 64*32 + 32*16 + 16*1)
            time_complexity.append(ops)
        
        # Get complexity metrics
        metrics = self.nlg_engine.calculate_complexity_metrics(
            self.numeric_data.shape[1]
        )
        
        explanation = f"""
[COMPLEXITY] Algorithm Complexity Analysis:

[SPACE] Space Complexity: {metrics['space_complexity']}
[PARAMETERS] Total Neural Network Parameters: {metrics['total_parameters']:,}
[NLG_PARAMS] NLG Model Parameters: {metrics['nlg_model_parameters']:,}

[TIME] Time Complexity: {metrics['time_complexity']}
[OPERATIONS] Estimated Operations per Training: {metrics['estimated_operations']:,}

[BIG_O] Big-O Analysis:
• Space: O(n*m) - Linear with input dimensions and hidden units
• Time: O(epochs * batch_size * parameters) - Proportional to training iterations

[SCALABILITY] The algorithm scales efficiently up to 500+ input dimensions
with polynomial time complexity during training.
        """
        
        return explanation
    
    def _provide_general_summary(self):
        """Provide general summary of the data using NLG"""
        nlg_data = {
            'total_regions': len(self.data),
            'key_metrics': 'segregation rates, urbanization levels, and organic waste composition'
        }
        
        summary = [
            f"Total regions analyzed: {len(self.data)}",
            f"Average segregation rate: {self.numeric_data['sor'].mean():.1f}%",
            f"Average urbanization: {self.numeric_data['urb'].mean():.1f}%",
            f"Average organic waste: {self.numeric_data['organic'].mean():.1f}%",
            "Data covers waste density, collection rates, and policy indicators"
        ]
        
        nlg_response = self.nlg_engine.generate_data_to_text(nlg_data, "summary")
        
        return f"[SUMMARY] Data Overview:\n\n{nlg_response}\n\n[STATISTICS] Key Metrics:\n" + "\n".join(f"• {item}" for item in summary)
    
    def visualize_complexity(self):
        """Create comprehensive complexity visualization"""
        input_dims = [10, 20, 50, 100, 200, 500]
        
        # Calculate complexities
        space_complexity = []
        time_complexity = []
        
        for dim in input_dims:
            params = dim * 64 + 64 * 32 + 32 * 16 + 16 * 1
            space_complexity.append(params)
            
            ops = 50 * 32 * (dim * 64 + 64*32 + 32*16 + 16*1)
            time_complexity.append(ops)
        
        # Create figure with subplots
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(15, 12))
        
        # Plot 1: Space Complexity
        ax1.plot(input_dims, space_complexity, 'b-o', linewidth=2, markersize=8)
        ax1.fill_between(input_dims, space_complexity, alpha=0.3)
        ax1.set_title('Space Complexity: O(n*m)', fontsize=14, fontweight='bold')
        ax1.set_xlabel('Input Dimension (n)', fontsize=12)
        ax1.set_ylabel('Number of Parameters', fontsize=12)
        ax1.grid(True, alpha=0.3)
        ax1.set_yscale('log')
        
        # Plot 2: Time Complexity
        ax2.plot(input_dims, time_complexity, 'r-s', linewidth=2, markersize=8)
        ax2.fill_between(input_dims, time_complexity, alpha=0.3, color='red')
        ax2.set_title('Time Complexity: O(e*b*ops)', fontsize=14, fontweight='bold')
        ax2.set_xlabel('Input Dimension (n)', fontsize=12)
        ax2.set_ylabel('Estimated Operations', fontsize=12)
        ax2.grid(True, alpha=0.3)
        ax2.set_yscale('log')
        
        # Plot 3: Parameter Growth Rate
        growth_rates = [space_complexity[i] / max(1, space_complexity[i-1]) 
                       for i in range(1, len(space_complexity))]
        ax3.bar(range(len(growth_rates)), growth_rates, color='green', alpha=0.7)
        ax3.set_title('Parameter Growth Rate (Sequential)', fontsize=14, fontweight='bold')
        ax3.set_xlabel('Transition Index', fontsize=12)
        ax3.set_ylabel('Growth Factor', fontsize=12)
        ax3.grid(True, alpha=0.3, axis='y')
        
        # Plot 4: Scalability Analysis
        scalability = [t/s for s, t in zip(space_complexity, time_complexity)]
        ax4.plot(input_dims, scalability, 'g-^', linewidth=2, markersize=8)
        ax4.set_title('Algorithm Efficiency (Time/Space Ratio)', fontsize=14, fontweight='bold')
        ax4.set_xlabel('Input Dimension (n)', fontsize=12)
        ax4.set_ylabel('Efficiency Ratio', fontsize=12)
        ax4.grid(True, alpha=0.3)
        
        plt.tight_layout()
        plt.savefig('complexity_visualization.png', dpi=300, bbox_inches='tight')
        print("\n[SAVED] Complexity visualization saved as 'complexity_visualization.png'")
        print("[DISPLAY] Rendering visualization...")
        plt.show()
        
        return "[SUCCESS] Complexity visualization complete and saved."


# ============================================================================
# MAIN EXECUTION - INTERACTIVE MODE
# ============================================================================
def main():
    """Execute policy analysis pipeline"""
    
    print("\n" + "="*70)
    print("WASTE MANAGEMENT POLICYMAKING ASSISTANT")
    print("Interactive Policy Analysis System")
    print("="*70)
    print("\n[FEATURES] System Components:")
    print("  * Neural Network: 4-layer MLP with batch normalization")
    print("  * NLG Engine: BART-large-cnn (facebook/bart-large-cnn)")
    print("  * Data Visualization: Matplotlib complexity analysis")
    print("  * Extractive Approach: No hallucination, data-grounded responses")
    print("="*70 + "\n")
    
    # Initialize
    config = WasteManagementPolicyConfig()
    config.display_config()
    
    # Data preprocessing
    print("[STAGE-1] DATA PREPARATION")
    print("-" * 70)
    preprocessor = PolicyDataPreprocessor(config)
    df = preprocessor.get_processed_data()
    print(f"\n[SUCCESS] Data prepared: {df.shape[0]} regions, {df.shape[1]} indicators\n")
    
    # Model training
    print("[STAGE-2] MODEL TRAINING")
    print("-" * 70)
    
    numeric_df = df.select_dtypes(include='number')
    X = numeric_df.values
    y = (numeric_df.iloc[:, 0] > numeric_df.iloc[:, 0].median()).astype(int).values
    
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=config.test_size, random_state=config.random_state
    )
    X_train, X_val, y_train, y_val = train_test_split(
        X_train, y_train, test_size=0.2, random_state=config.random_state
    )
    
    generator = PolicyIntelligenceGenerator(config)
    generator.build_model(input_dim=X_train.shape[1])
    history = generator.train_model(X_train, y_train, X_val, y_val)
    
    # Evaluation
    print("\n[STAGE-3] POLICY ANALYSIS")
    print("-" * 70)
    
    y_pred = generator.model.predict(X_test)
    accuracy = accuracy_score(y_test, y_pred)
    print(f"[EVALUATION] Model accuracy: {accuracy:.2%}\n")
    
    # Sample policy brief
    print("[STAGE-4] SAMPLE POLICY BRIEF")
    print("-" * 70)
    
    if len(df) > 0:
        sample_region = df.iloc[0].to_dict()
        sample_region['region'] = 'Metro Manila'
        sample_region['province'] = 'NCR'
        
        policy_brief = generator.generate_policy_brief(sample_region)
        evaluation = generator.evaluate_region(sample_region)
        
        print(f"\nRegion: {policy_brief['region']}")
        print(f"Assessment Date: {policy_brief['assessment_date']}")
        print(f"\nWaste Management Maturity Score: {evaluation['waste_management_score']:.1f}/100")
        print(f"Policy Readiness: {evaluation['policy_readiness']}")
        print(f"\nKey Recommendations: {len(policy_brief['policy_recommendations'])} policies identified")
    
    print("\n" + "="*70)
    print("ANALYSIS COMPLETE - System ready for interactive use")
    print("="*70 + "\n")
    
    return {
        'preprocessor': preprocessor,
        'generator': generator,
        'model': generator.model,
        'data': df,
        'metrics_recorder': metrics_recorder
    }


def interactive_mode(results):
    """Start interactive question-answering mode with NLG"""
    qa_system = PolicyQuestionAnswering(results['data'], results['generator'])
    metrics_recorder = results.get('metrics_recorder')
    
    print("\n" + "="*70)
    print("INTERACTIVE POLICY ASSISTANT - ENHANCED WITH NLG & BART")
    print("="*70)
    print("\n[AVAILABLE] Commands and Options:")
    print("  • Ask policy questions about waste management (extractive NLG responses)")
    print("  • Type 'complexity' or 'algorithm' to analyze complexity")
    print("  • Type 'visualize' to display space/time complexity graphs")
    print("  • Type 'metrics' to view algorithm performance metrics")
    print("  • Type 'export' or 'excel' to export metrics to Excel file")
    print("  • Type 'help' for detailed topic list")
    print("  • Type 'quit', 'exit', or 'q' to end session\n")
    print("="*70 + "\n")
    
    while True:
        try:
            question = input("Your question: ").strip()
            
            if question.lower() in ['quit', 'exit', 'q']:
                print("\n[EXIT] Thank you for using the Waste Management Policy Assistant!")
                break
            
            if question.lower() in ['help', '?']:
                print("\n[HELP] Available Query Topics:")
                print("  • 'improvements needed?' - What areas require improvement")
                print("  • 'worst regions?' - Priority regions requiring attention")
                print("  • 'segregation issues?' - Waste segregation analysis")
                print("  • 'organic waste?' - Organic waste management strategies")
                print("  • 'urban challenges?' - Urban waste management concerns")
                print("  • 'maturity score?' - Policy readiness assessment")
                print("  • 'policy recommendations?' - Evidence-based policy options")
                print("  • 'complexity analysis?' - Algorithm efficiency details")
                print("  • 'visualize' - Display complexity visualization")
                print("  • 'metrics' - View algorithm performance metrics")
                print("  • 'export' or 'excel' - Export metrics to Excel file\n")
                continue
            
            if question.lower() in ['export', 'excel']:
                if metrics_recorder:
                    print("\n[EXPORTING] Saving metrics to Excel file...")
                    success = metrics_recorder.save_to_excel('algorithm_metrics.xlsx')
                    if success:
                        print("[SUCCESS] Excel file created: algorithm_metrics.xlsx\n")
                    else:
                        print("[ERROR] Failed to create Excel file.\n")
                else:
                    print("\n[ERROR] Metrics recorder not available.\n")
                continue
            
            if question.lower() == 'metrics':
                if metrics_recorder:
                    latest = metrics_recorder.get_latest_metrics()
                    if latest:
                        metrics_recorder.print_metrics_report(latest)
                    else:
                        print("\n[INFO] No metrics recorded yet.\n")
                else:
                    print("\n[ERROR] Metrics recorder not available.\n")
                continue
            
            if question.lower() == 'visualize':
                print("\n[GENERATING] Creating complexity visualization...")
                qa_system.visualize_complexity()
                continue
            
            if not question:
                continue
            
            # Use extractive NLG to answer question
            answer = qa_system.answer_question(question)
            print(f"\n{answer}\n")
            print("-" * 70)
            
        except KeyboardInterrupt:
            print("\n\n[TERMINATED] Session ended by user.")
            break
        except Exception as e:
            print(f"[ERROR] Issue processing question: {e}")
            continue


if __name__ == "__main__":
    results = main()
    interactive_mode(results)
